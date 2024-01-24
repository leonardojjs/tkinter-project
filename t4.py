from tkinter import *
from openpyxl.workbook import workbook
from openpyxl import load_workbook

root = Tk()
root.title("Aplikasi Olah Excel")
root.iconbitmap("iconmars.ico")
root.geometry("500x500")

def select(e):
    my_label.config(text=my_listbox.get(ANCHOR))

my_listbox = Listbox(root, width=45)
my_listbox.pack(pady=20)

wb = load_workbook("Fruits.xlsx")
ws = wb.active

col_a = ws["A"]
col_b = ws["B"]

for item in col_a:
    my_listbox.insert(END, item.value)

my_label = Label(root, text="Select item")
my_label.pack(pady=20)
my_listbox.bind("<ButtonRelease-1>", select)


root.mainloop()